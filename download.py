import openpyxl
from openpyxl import load_workbook
import time
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import re
import urllib.request # ライブラリを取り込む
import socket


import sys

args = sys.argv

pattern = '内国株式'

wb = load_workbook(filename = 'data_j.xlsx')
ws = wb["Sheet1"]
rg = ws["B2:" + "D" + str(ws.max_row)]
for row in rg:
    if len(args) == 2 and row[0].value < int(args[1]):
        continue;
    if re.search(pattern, row[2].value):
        print(row[2].value)
        print(row[1].value)

        #googleで検索する文字
        search_string = row[1].value + ' 決算説明資料 filetype:pdf'

        #Seleniumを使うための設定とgoogleの画面への遷移
        INTERVAL = 2.5
        URL = "https://www.google.com/"
        driver_path = "./chromedriver"
        driver = webdriver.Chrome(executable_path=driver_path)
        driver.maximize_window()
        time.sleep(INTERVAL)
        driver.get(URL)
        time.sleep(INTERVAL)

        #文字を入力して検索
        driver.find_element_by_name('q').send_keys(search_string)
        driver.find_elements_by_name('btnK')[1].click() #btnKが2つあるので、その内の後の方
        time.sleep(INTERVAL)

        #検索結果の一覧を取得する
        results = []
        flag = False
        while True:
            g_ary = driver.find_elements_by_class_name('yuRUbf')
            for g in g_ary:
                url = g.find_element_by_tag_name('a').get_attribute('href')
                print (url)
                if re.search("pdf$", url):

                    try:
                        opener=urllib.request.build_opener()
                        opener.addheaders=[('User-Agent','Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1941.0 Safari/537.36')]
                        urllib.request.install_opener(opener)
                        socket.setdefaulttimeout(15)
                        urllib.request.urlretrieve(url, "./pdf/" + str(row[0].value) + '_' + str(row[1].value) + '.pdf')
                        flag = True
                        break
                    except urllib.error.HTTPError as e:
                        print('raise HTTPError')
                        print(e.code)
                        print(e.reason)
                        continue
                    except urllib.error.URLError as e:
                        print('raise URLError')
                        print(e.reason)
                        continue
                    except socket.timeout as e:
                        print ("socket timeout")

            if flag:
                break
            try:
                driver.find_element_by_id('pnnext').click()
                time.sleep(INTERVAL)
            except NoSuchElementException:
                flag = True
                break
        driver.close()


